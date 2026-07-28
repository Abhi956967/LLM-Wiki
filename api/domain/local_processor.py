"""Local document processor — runs extraction without S3 or Postgres.

Processes files directly from the workspace filesystem and updates SQLite.
Respects PDF_BACKEND config and optional Mistral/LibreOffice backends.
"""

import asyncio
import contextvars
import functools
import json
import logging
import os
import shutil
import signal
import subprocess
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path

import aiosqlite
from config import settings
from domain.file_types import (
    HTML_TYPES,
    IMAGE_TYPES,
    OFFICE_TYPES,
    PDF_TYPES,
    PROCESSING_TYPES,
    SIMPLE_TEXT_TYPES,
    SPREADSHEET_TYPES,
)
from domain.watcher import mark_written
from infra.db.sqlite import SQLiteDocumentRepository, create_pool
from services.extracted_assets import build_pdf_image_assets

logger = logging.getLogger(__name__)

# Cap concurrent fire-and-forget extractions so a burst of dropped files can't
# spawn one LibreOffice/OCR job (and connection) per file at once.
PROCESS_CONCURRENCY = 4
_process_semaphore = asyncio.Semaphore(PROCESS_CONCURRENCY)


async def _to_thread_joined(func, /, *args, **kwargs):
    """Run blocking work in a thread and join it before propagating cancellation.

    `asyncio.to_thread()` cancels only its asyncio waiter, not the underlying
    thread. This helper keeps the executor future alive under cancellation and
    waits for the real worker to finish before cleanup/status transitions run.
    """
    loop = asyncio.get_running_loop()
    context = contextvars.copy_context()
    call = functools.partial(context.run, func, *args, **kwargs)
    worker = loop.run_in_executor(None, call)

    try:
        return await asyncio.shield(worker)
    except asyncio.CancelledError:
        # A shutdown may send more than one cancellation. Keep shielding until
        # the executor future is actually done; it is a Future rather than an
        # asyncio Task, so loop-wide task cancellation cannot cancel it behind
        # our back while its thread continues running.
        while not worker.done():
            try:
                await asyncio.shield(worker)
            except asyncio.CancelledError:
                continue
            except BaseException:
                break

        # Retrieve a post-cancellation worker exception so it is not reported
        # as unobserved. The request/task cancellation remains authoritative.
        if worker.done():
            try:
                worker.result()
            except BaseException:
                pass
        raise


def _write_bytes_file(path: Path, content: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)


def _copy_file(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)


def _find_libreoffice() -> str | None:
    return shutil.which("libreoffice") or shutil.which("soffice")


def _run_process_group(command: list[str], timeout: int) -> subprocess.CompletedProcess:
    """Run a conversion in a process group so timeout cannot orphan children."""
    popen_options: dict = {
        "stdout": subprocess.PIPE,
        "stderr": subprocess.PIPE,
    }
    if os.name == "posix":
        popen_options["start_new_session"] = True
    elif hasattr(subprocess, "CREATE_NEW_PROCESS_GROUP"):  # pragma: no cover - Windows
        popen_options["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP

    proc = subprocess.Popen(command, **popen_options)
    try:
        stdout, stderr = proc.communicate(timeout=timeout)
    except subprocess.TimeoutExpired as error:
        if proc.poll() is None:
            if os.name == "posix":
                try:
                    os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
                except ProcessLookupError:
                    pass
            else:  # pragma: no cover - Windows
                subprocess.run(
                    ["taskkill", "/F", "/T", "/PID", str(proc.pid)],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    check=False,
                )
                if proc.poll() is None:
                    proc.kill()
        proc.wait()
        raise TimeoutError(f"Conversion timed out after {timeout} seconds") from error

    return subprocess.CompletedProcess(command, proc.returncode, stdout, stderr)


def _list_pdf_files(directory: Path) -> list[Path]:
    return list(directory.glob("*.pdf"))


def _encode_file_base64(path: Path) -> str:
    import base64

    return base64.b64encode(path.read_bytes()).decode()


def _extract_spreadsheet_content(file_path: Path) -> list[tuple[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(file_path), read_only=True, data_only=True)
    try:
        sheets: list[tuple[str, str]] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = [
                " | ".join(str(cell) if cell is not None else "" for cell in row)
                for row in worksheet.iter_rows(values_only=True)
            ]
            sheets.append((sheet_name, "\n".join(rows)))
        return sheets
    finally:
        workbook.close()


def _extract_html_content(file_path: Path) -> str:
    raw_html = file_path.read_text(encoding="utf-8", errors="replace")

    try:
        from html_parser import Parser

        parser = Parser(raw_html, content_only=True)
        return parser.parse().content
    except Exception:
        return raw_html


async def process_document(db: aiosqlite.Connection, doc_id: str, workspace: Path) -> None:
    """Atomically claim a pending document, then extract text, chunk, update index."""
    claim = await db.execute(
        "UPDATE documents SET status = 'processing', error_message = NULL, "
        "updated_at = datetime('now') WHERE id = ? AND status = 'pending'",
        (doc_id,),
    )
    try:
        await db.commit()
    except asyncio.CancelledError:
        if claim.rowcount != 0:
            try:
                await _restore_cancelled_claim(db, doc_id)
            except Exception:
                logger.exception("Failed to restore cancelled document %s", doc_id[:8])
        raise
    if claim.rowcount == 0:
        return

    filename = doc_id[:8]
    try:
        cursor = await db.execute(
            "SELECT filename, file_type, relative_path FROM documents WHERE id = ?",
            (doc_id,),
        )
        row = await cursor.fetchone()
        if not row:
            logger.warning("Document %s not found", doc_id[:8])
            return

        cols = [d[0] for d in cursor.description]
        doc = dict(zip(cols, row))
        filename = doc["filename"]

        file_type = doc["file_type"] or ""
        file_path = workspace / doc["relative_path"]

        if not await _to_thread_joined(file_path.is_file):
            await db.execute(
                "UPDATE documents SET status = 'failed', error_message = 'File not found', "
                "updated_at = datetime('now') WHERE id = ?",
                (doc_id,),
            )
            await db.commit()
            return

        if file_type in PDF_TYPES:
            await _process_pdf(db, doc_id, file_path, workspace)
        elif file_type in OFFICE_TYPES:
            await _process_office(db, doc_id, file_path, workspace)
        elif file_type in SPREADSHEET_TYPES:
            await _process_spreadsheet(db, doc_id, file_path)
        elif file_type in IMAGE_TYPES:
            await _process_image(db, doc_id)
        elif file_type in HTML_TYPES:
            await _process_html(db, doc_id, file_path)
        else:
            await db.execute(
                "UPDATE documents SET status = 'ready', updated_at = datetime('now') WHERE id = ?",
                (doc_id,),
            )
            await db.commit()

        logger.info("Processed %s: %s", filename, file_type)

    except asyncio.CancelledError:
        # A shutdown can cancel extraction after the pending -> processing
        # claim. Make the document retryable before propagating cancellation.
        try:
            await _restore_cancelled_claim(db, doc_id)
        except Exception:
            logger.exception("Failed to restore cancelled document %s", doc_id[:8])
        raise
    except Exception as e:
        error_msg = str(e)[:500]
        try:
            await db.execute(
                "UPDATE documents SET status = 'failed', error_message = ?, "
                "updated_at = datetime('now') WHERE id = ?",
                (error_msg, doc_id),
            )
            await db.commit()
        except Exception:
            logger.exception("Failed to persist extraction failure for %s", doc_id[:8])
        logger.error("Failed to process %s: %s", filename, e)


async def _reset_processing_to_pending(db: aiosqlite.Connection, doc_id: str) -> None:
    await db.execute(
        "UPDATE documents SET status = 'pending', error_message = NULL, "
        "updated_at = datetime('now') WHERE id = ? AND status = 'processing'",
        (doc_id,),
    )
    await db.commit()


async def _restore_cancelled_claim(db: aiosqlite.Connection, doc_id: str) -> None:
    """Finish status cleanup even if shutdown sends a second cancellation."""
    cleanup = asyncio.create_task(_reset_processing_to_pending(db, doc_id))
    try:
        await asyncio.shield(cleanup)
    except asyncio.CancelledError:
        await cleanup


async def process_document_isolated(workspace: Path, doc_id: str) -> None:
    """Process a document on its own connection so fire-and-forget tasks can't
    flush another writer's open transaction on a shared connection."""
    try:
        async with _process_semaphore:
            db = await create_pool(str(workspace / ".llmwiki" / "index.db"), init_schema=False)
            try:
                await process_document(db, doc_id, workspace)
            finally:
                await db.close()
    except asyncio.CancelledError:
        raise
    except Exception:
        # This coroutine is intentionally spawned as a background task. Keep
        # unexpected connection/setup failures from becoming unobserved task
        # exceptions; pending documents remain eligible for reconciliation.
        logger.exception("Isolated processing task failed for %s", doc_id[:8])


async def chunk_text_document(db: aiosqlite.Connection, doc_id: str, content: str | None) -> None:
    """Chunk an already-extracted text document so it becomes full-text searchable."""
    from services.chunker import chunk_text

    chunks = await _to_thread_joined(chunk_text, content or "")
    await _store_chunks(db, doc_id, chunks)
    # `parser` doubles as the chunked-marker so reconcile skips docs that
    # legitimately produce zero chunks (empty/short) instead of retrying them.
    await db.execute(
        "UPDATE documents SET parser = 'text', updated_at = datetime('now') WHERE id = ?",
        (doc_id,),
    )
    await db.commit()


def reconcile_workspace(db: aiosqlite.Connection, workspace: Path):
    """Process documents that were indexed but never extracted or chunked.

    `llmwiki init` lists existing files into the index without extracting PDFs
    or building search chunks; this backfills both so a folder pointed at on
    first run is actually readable and searchable.
    """
    # This function intentionally captures the cutoff before returning its
    # coroutine. Main creates the reconcile task before the watcher task, so
    # extractions claimed after startup cannot be mistaken for stale work.
    recovery_cutoff = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    return _reconcile_workspace(db, workspace, recovery_cutoff)


async def _reconcile_workspace(
    db: aiosqlite.Connection,
    workspace: Path,
    recovery_cutoff: str,
) -> None:
    interrupted_ids = await _recover_interrupted_documents(db, recovery_cutoff)
    unchunked_ids = await _unchunked_extractable_ids(db)
    extract_ids = list(dict.fromkeys([*interrupted_ids, *unchunked_ids]))
    for doc_id in extract_ids:
        try:
            await db.execute(
                "UPDATE documents SET status = 'pending', updated_at = datetime('now') WHERE id = ?",
                (doc_id,),
            )
            await db.commit()
            await process_document(db, doc_id, workspace)
        except Exception:
            logger.exception("Reconcile: failed to process %s", doc_id[:8])

    text_docs = await _unchunked_text_docs(db)
    for doc_id, content in text_docs:
        try:
            await chunk_text_document(db, doc_id, content)
        except Exception:
            logger.exception("Reconcile: failed to chunk %s", doc_id[:8])

    if extract_ids or text_docs:
        logger.info(
            "Reconciled workspace: %d extracted, %d text-chunked",
            len(extract_ids), len(text_docs),
        )


async def _store_chunks(db: aiosqlite.Connection, doc_id: str, chunks: list) -> None:
    """Store chunks into SQLite, replacing any existing ones."""
    await db.execute("DELETE FROM document_chunks WHERE document_id = ?", (doc_id,))
    for c in chunks:
        await db.execute(
            "INSERT INTO document_chunks (id, document_id, chunk_index, content, source_content, page, "
            "start_char, token_count, header_breadcrumb) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), doc_id, c.index, c.content, c.content, c.page,
             c.start_char, c.token_count, c.header_breadcrumb),
        )


# ── PDF extraction ────────────────────────────────────────────────────────

async def _save_local_images(
    db: aiosqlite.Connection, doc_id: str, workspace: Path,
    pages_with_images: list[tuple[int, str, list[dict]]],
) -> dict[int, dict]:
    """Save extracted images as hidden sibling assets and return page metadata."""
    repo = SQLiteDocumentRepository(db)
    doc = await repo.get(doc_id)
    if not doc:
        return {}

    assets, page_elements = await _to_thread_joined(
        build_pdf_image_assets,
        doc_id,
        doc["filename"],
        doc["path"],
        pages_with_images,
    )
    if not assets:
        return {}

    await db.execute(
        "DELETE FROM documents WHERE source_kind = 'asset' AND metadata LIKE ?",
        (f'%"parent_document_id": "{doc_id}"%',),
    )
    await db.commit()

    asset_metadata = []
    for asset in assets:
        relative_asset = (asset.path.rstrip("/") + "/" + asset.filename).lstrip("/")
        local_asset = workspace / relative_asset
        mark_written(str(local_asset))
        await _to_thread_joined(_write_bytes_file, local_asset, asset.data)
        await repo.create_asset(
            asset.document_id,
            doc["user_id"],
            asset.filename,
            asset.path,
            asset.filename,
            asset.file_type,
            len(asset.data),
            asset.metadata(),
        )
        asset_metadata.append(asset.metadata())

    await repo.set_metadata_field(doc_id, "assets", asset_metadata)
    return page_elements


async def _store_page_contents(
    db: aiosqlite.Connection, doc_id: str,
    page_contents: list[tuple[int, str]], parser: str,
    page_elements: dict[int, dict] | None = None,
) -> None:
    """Store extracted pages, chunks, and update document status."""
    num_pages = len(page_contents)

    await db.execute("DELETE FROM document_pages WHERE document_id = ?", (doc_id,))
    for page_num, content in page_contents:
        elements = (page_elements or {}).get(page_num)
        await db.execute(
            "INSERT INTO document_pages (id, document_id, page, content, elements) VALUES (?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), doc_id, page_num, content,
             json.dumps(elements) if elements else None),
        )

    full_content = "\n\n---\n\n".join(md for _, md in page_contents)

    from services.chunker import chunk_pages
    chunks = await _to_thread_joined(chunk_pages, page_contents)
    await _store_chunks(db, doc_id, chunks)

    await db.execute(
        "UPDATE documents SET status = 'ready', content = ?, page_count = ?, "
        "parser = ?, updated_at = datetime('now') WHERE id = ?",
        (full_content, num_pages, parser, doc_id),
    )
    await db.commit()


async def _process_pdf(db: aiosqlite.Connection, doc_id: str, file_path: Path, workspace: Path) -> None:
    """Extract PDF text. Uses opendataloader by default, Mistral if configured."""
    if settings.PDF_BACKEND == "mistral" and settings.MISTRAL_API_KEY:
        await _process_pdf_mistral(db, doc_id, file_path, workspace)
    else:
        from services.pdf_extract import extract_pdf
        pages_with_images = await _to_thread_joined(extract_pdf, str(file_path))
        page_elements = await _save_local_images(db, doc_id, workspace, pages_with_images)
        page_contents = [(num, md) for num, md, _ in pages_with_images]
        await _store_page_contents(db, doc_id, page_contents, "opendataloader", page_elements)


# ── Office processing ─────────────────────────────────────────────────────

async def _process_office(db: aiosqlite.Connection, doc_id: str, file_path: Path, workspace: Path) -> None:
    """Convert Office docs to PDF via local LibreOffice, then extract text."""
    lo = await _to_thread_joined(_find_libreoffice)
    if not lo:
        await db.execute(
            "UPDATE documents SET status = 'failed', "
            "error_message = 'LibreOffice not installed. Install it to process Office files.', "
            "updated_at = datetime('now') WHERE id = ?",
            (doc_id,),
        )
        await db.commit()
        return

    tmpdir = Path(await _to_thread_joined(tempfile.mkdtemp))
    try:
        result = await _to_thread_joined(
            _run_process_group,
            [
                lo,
                f"-env:UserInstallation=file://{tmpdir}/lo-profile",
                "--headless",
                "--norestore",
                "--nofirststartwizard",
                "--convert-to",
                "pdf",
                "--outdir",
                str(tmpdir),
                str(file_path),
            ],
            120,
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice conversion failed: {result.stderr.decode()[:300]}")

        pdf_files = await _to_thread_joined(_list_pdf_files, tmpdir)
        if not pdf_files:
            raise RuntimeError("LibreOffice produced no PDF output")

        converted_pdf = pdf_files[0]

        # Store converted PDF in cache for the viewer
        cache_dir = workspace / ".llmwiki" / "cache" / "local" / doc_id
        await _to_thread_joined(
            _copy_file,
            converted_pdf,
            cache_dir / "converted.pdf",
        )

        from services.pdf_extract import extract_pdf
        pages_with_images = await _to_thread_joined(extract_pdf, str(converted_pdf))
        page_elements = await _save_local_images(db, doc_id, workspace, pages_with_images)
        page_contents = [(num, md) for num, md, _ in pages_with_images]
        await _store_page_contents(db, doc_id, page_contents, "libreoffice+opendataloader", page_elements)
    finally:
        await _to_thread_joined(shutil.rmtree, tmpdir, True)


# ── Mistral OCR ───────────────────────────────────────────────────────────

async def _process_pdf_mistral(db: aiosqlite.Connection, doc_id: str, file_path: Path, workspace: Path) -> None:
    """Extract PDF via Mistral OCR API (better tables/layout, requires API key)."""
    import httpx

    pdf_b64 = await _to_thread_joined(_encode_file_base64, file_path)

    async with httpx.AsyncClient(timeout=120) as client:
        resp = await client.post(
            "https://api.mistral.ai/v1/ocr",
            headers={"Authorization": f"Bearer {settings.MISTRAL_API_KEY}"},
            json={
                "model": "mistral-ocr-latest",
                "document": {"type": "document_url", "document_url": f"data:application/pdf;base64,{pdf_b64}"},
            },
        )
        resp.raise_for_status()
        result = resp.json()

    pages = result.get("pages", [])
    page_contents = [(i + 1, p.get("markdown", "")) for i, p in enumerate(pages)]
    await _store_page_contents(db, doc_id, page_contents, "mistral")


# ── Spreadsheet processing ────────────────────────────────────────────────

async def _process_spreadsheet(db: aiosqlite.Connection, doc_id: str, file_path: Path) -> None:
    """Extract spreadsheet data via openpyxl. Stores pages AND chunks for search."""
    sheets = await _to_thread_joined(_extract_spreadsheet_content, file_path)

    await db.execute("DELETE FROM document_pages WHERE document_id = ?", (doc_id,))

    all_content = []
    page_contents = []
    for i, (sheet_name, content) in enumerate(sheets, 1):
        elements = json.dumps({"sheet_name": sheet_name})

        await db.execute(
            "INSERT INTO document_pages (id, document_id, page, content, elements) "
            "VALUES (?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), doc_id, i, content, elements),
        )
        all_content.append(f"## {sheet_name}\n\n{content}")
        page_contents.append((i, content))

    num_sheets = len(sheets)
    full_content = "\n\n".join(all_content)

    from services.chunker import chunk_pages
    chunks = await _to_thread_joined(chunk_pages, page_contents)
    await _store_chunks(db, doc_id, chunks)

    await db.execute(
        "UPDATE documents SET status = 'ready', content = ?, page_count = ?, "
        "parser = 'openpyxl', updated_at = datetime('now') WHERE id = ?",
        (full_content, num_sheets, doc_id),
    )
    await db.commit()


# ── Image / HTML processing ──────────────────────────────────────────────

async def _process_image(db: aiosqlite.Connection, doc_id: str) -> None:
    """Images are stored as-is — just mark ready."""
    await db.execute(
        "UPDATE documents SET status = 'ready', page_count = 1, "
        "parser = 'native', updated_at = datetime('now') WHERE id = ?",
        (doc_id,),
    )
    await db.commit()


async def _process_html(db: aiosqlite.Connection, doc_id: str, file_path: Path) -> None:
    """Extract HTML content via webmd parser."""
    content = await _to_thread_joined(_extract_html_content, file_path)

    from services.chunker import chunk_text
    chunks = await _to_thread_joined(chunk_text, content)
    await _store_chunks(db, doc_id, chunks)

    await db.execute(
        "UPDATE documents SET status = 'ready', content = ?, page_count = 1, "
        "parser = 'webmd', updated_at = datetime('now') WHERE id = ?",
        (content, doc_id),
    )
    await db.commit()


# ── Reconciliation queries ────────────────────────────────────────────────

async def _recover_interrupted_documents(
    db: aiosqlite.Connection,
    recovery_cutoff: str,
) -> list[str]:
    """Restore documents left processing by a prior process interruption."""
    placeholders = ",".join("?" for _ in PROCESSING_TYPES)
    params = tuple(PROCESSING_TYPES)
    cursor = await db.execute(
        f"SELECT id FROM documents WHERE status = 'processing' AND source_kind != 'asset' "
        f"AND file_type IN ({placeholders}) "
        f"AND (updated_at IS NULL OR updated_at < ?)",
        (*params, recovery_cutoff),
    )
    document_ids = [row[0] for row in await cursor.fetchall()]
    if not document_ids:
        return []

    await db.execute(
        f"UPDATE documents SET status = 'pending', error_message = NULL, "
        f"updated_at = datetime('now') WHERE status = 'processing' "
        f"AND source_kind != 'asset' AND file_type IN ({placeholders}) "
        f"AND (updated_at IS NULL OR updated_at < ?)",
        (*params, recovery_cutoff),
    )
    await db.commit()
    logger.warning(
        "Recovered %d interrupted document extraction(s)",
        len(document_ids),
    )
    return document_ids


async def _unchunked_extractable_ids(db: aiosqlite.Connection) -> list[str]:
    """IDs of documents still needing local background processing.

    Excludes 'processing' so reconcile never reclaims a doc an isolated task is mid-extracting.
    """
    placeholders = ",".join("?" for _ in PROCESSING_TYPES)
    cursor = await db.execute(
        f"SELECT id FROM documents WHERE status NOT IN ('failed', 'processing') AND source_kind != 'asset' "
        f"AND parser IS NULL "
        f"AND file_type IN ({placeholders}) "
        f"AND id NOT IN (SELECT DISTINCT document_id FROM document_chunks)",
        tuple(PROCESSING_TYPES),
    )
    return [r[0] for r in await cursor.fetchall()]


async def _unchunked_text_docs(db: aiosqlite.Connection) -> list[tuple[str, str]]:
    """(id, content) for never-chunked simple-text docs that have content."""
    placeholders = ",".join("?" for _ in SIMPLE_TEXT_TYPES)
    cursor = await db.execute(
        f"SELECT id, content FROM documents WHERE status NOT IN ('failed', 'processing') AND source_kind != 'asset' "
        f"AND parser IS NULL "
        f"AND file_type IN ({placeholders}) "
        f"AND content IS NOT NULL AND content != '' "
        f"AND id NOT IN (SELECT DISTINCT document_id FROM document_chunks)",
        tuple(SIMPLE_TEXT_TYPES),
    )
    return [(r[0], r[1]) for r in await cursor.fetchall()]
